#•	Byrja lesa alvöru gögn.
#o	Main fall í python sem les skjal, ítrar worksheets í fall x
#o	X les dálka, ítrar línur í gögn
#o	Look ahead til að finna enda á fjölskyldur
#o	Sendir gögn í fall y
#o	Y les dálk og parsar, les persónur og parsar

#Þetta les excel skjal og færir gögnin yfir í gagnagrunn

from openpyxl import load_workbook, worksheet
from openpyxl.styles import Font #sjá bold=True
wb = load_workbook(filename='Vinnuskjal1.xlsx')
ws = wb['Sheet1']

import sqlite3 as lite
#import sys
con = lite.connect('data.db')
cur = con.cursor()
db_fields = ['Id', 'num', 'Idnum']

for row in ws.get_squared_range(ws.min_column, ws.min_row, ws.max_column, ws.max_row):
    placeholders = []
    vals = []

    for cell in row:
        placeholders.append('?')
        vals.append(cell.value)
    sql = 'INSERT INTO tilraun (' + ','.join(db_fields) + ') VALUES (' + ','.join(placeholders[:len(db_fields)]) + ')'
    with con:
        cur.execute(sql, tuple(vals[:len(db_fields)])) #make sure only db_fields number of columns

with con:
    cur.execute("SELECT * FROM tilraun")
    rows = cur.fetchall()
    for row in rows:
        print(row)
print("All done!")

#    data_sql = cur.fetchone()
#    while data_sql:
#        print(data_sql)
#        data_sql = cur.fetchone()
#    else:
    # print("Second done")

#SQL build for the data in the database

conn = lite.connect('data.db')
curr = conn.cursor()
with conn
    curr.execute('CREATE TABLE Area_names(area_ID int, hreppur varchar, farm varchar)')