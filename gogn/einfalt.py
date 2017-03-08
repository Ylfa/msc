#Þetta les excel skjal og færir gögnin yfir í gagnagrunn

##Skoða næsta skref - Flask

####################### EXCEL #################################
from openpyxl import load_workbook, worksheet
#from openpyxl.styles import Font #sjá bold=True

####################### SQLITE  ################################
import sqlite3 as lite
import sys
con = lite.connect('gogn.db')
cur = con.cursor()

def init_tables():
    print("Database start")
    with con:
        cur.execute('''DROP TABLE IF EXISTS FARM_NAMES''');
        cur.execute('''DROP TABLE IF EXISTS FARM_DATA''');
        cur.execute('''CREATE TABLE IF NOT EXISTS FARM_NAMES(
          ID INTEGER primary key autoincrement,
          AREA_NAME TEXT,
          FARM_NAME TEXT,
          YEAR_DATA BLOB,
          NAME_DATA BLOB);''')
    return("Database init done.")

def get_farm_col(list):
    """
    Takes a list of farms of an area and returns the indexes (col nr) of the farms and their names in tuples
    """
    data_list = []
    for index, elem in enumerate(list):
        data_list.append((index,elem))
    data_list = [data for data in data_list if data[1] not in (' ', '  ',None)]
    return data_list

def main():
    n = 0
    wb = load_workbook(filename='Vinnuskjal1.xlsx')
    ws = wb.worksheets[n]
    sheet_names = wb.get_sheet_names()

    a = init_tables()
    print(a)

    while len(sheet_names) > 1:
        data = [ws.cell(row=1, column=i).value for i in
                range(ws.min_column, ws.max_column + 2)]
        new_list = get_farm_col(data)
        data_cleaned = [item for item in data if item not in (None, ' ', '  ')]

    # Data to be stored in table farm_names:
        while len(data_cleaned) > 0:
            b = new_list[0][0]
            year_blob = [ws.cell(row=i, column=b).value for i in range(2, ws.max_row)]
            name_blob= [ws.cell(row=i, column=b+1).value for i in range(2, ws.max_row)]
            sql_insert = [ws.title, data_cleaned[0], str(year_blob), str(name_blob)]

            with con:
                cur.execute('INSERT INTO FARM_NAMES (AREA_NAME, FARM_NAME, YEAR_DATA, NAME_DATA) VALUES (?,?,?,?)', sql_insert)
                con.commit()
            del data_cleaned[0]
            del new_list[0]
        n = n + 1
        ws = wb.worksheets[n]
        sheet_names.pop(0)  # Popping because there was no function to get_max_sheets in file

main()

"""
with con:
    cur.execute("SELECT * FROM farm_names")
    rows = cur.fetchall()
    for row in rows:
        print(row)
print("All done!")
con.close()
"""
