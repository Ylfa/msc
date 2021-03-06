# Þetta les excel skjal og færir gögnin yfir í gagnagrunn

from openpyxl import load_workbook, worksheet
import json
import datetime
import sqlite3 as lite
import sys
con = lite.connect('gogn.db')
cur = con.cursor()

def datetime_handler(x):
    if isinstance(x, datetime.datetime):
        return x.isoformat()
    raise TypeError("Unknown type")


def init_tables():
    print("Database start")
    with con:
        cur.execute('''DROP TABLE IF EXISTS AREA_ID''')
        cur.execute('''DROP TABLE IF EXISTS FARM_ID''')
        cur.execute('''DROP TABLE IF EXISTS FAMILY_ID''')
        cur.execute('''DROP TABLE IF EXISTS TIME_SPANS''')
        cur.execute('''DROP TABLE IF EXISTS SKAMMSTAFANIR''')
        cur.execute('''DROP TABLE IF EXISTS PERSON_ID''')
        cur.execute('''CREATE TABLE IF NOT EXISTS AREA_ID(
                  AREA_ID INTEGER primary key autoincrement,
                  AREA_NAME TEXT);''')

        cur.execute('''CREATE TABLE IF NOT EXISTS FARM_ID(
          AREA_ID INTEGER,
          FARM_ID INTEGER primary key autoincrement,
          AREA_NAME TEXT,
          FARM_NAME TEXT,
          FOREIGN KEY (AREA_ID) REFERENCES AREA_ID(AREA_ID));''')

        cur.execute('''CREATE TABLE IF NOT EXISTS FAMILY_ID(
                  FAMILY_ID INTEGER PRIMARY KEY AUTOINCREMENT,
                  FARM_ID INTEGER,
                  AREA_ID INTEGER,
                  AREA_NAME TEXT,
                  FARM_NAME TEXT,
                  FAMILY_YEAR TEXT,
                  FAMILY_DATA TEXT,
                  FOREIGN KEY (AREA_ID) REFERENCES AREA_ID(AREA_ID),
                  FOREIGN KEY (FARM_ID) REFERENCES FARM_ID(FARM_ID));''')
        cur.execute('''CREATE TABLE IF NOT EXISTS PERSON_ID(
                                  PERSON_ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                  FAMILY_ID INTEGER,
                                  FARM_ID INTEGER,
                                  AREA_ID INTEGER,
                                  AREA_NAME TEXT,
                                  FARM_NAME TEXT,
                                  FAMILY_YEAR TEXT,
                                  FAMILY_DATA TEXT,
                                  FOREIGN KEY (AREA_ID) REFERENCES AREA_ID(AREA_ID),
                                  FOREIGN KEY (FARM_ID) REFERENCES FARM_ID(FARM_ID),
                                  FOREIGN KEY (FAMILY_ID) REFERENCES FAMILY_ID(FAMILY_ID));''')
        cur.execute('''CREATE TABLE IF NOT EXISTS TIME_SPANS(
          FAMILY_ID INTEGER PRIMARY KEY,
          TIME_FROM INTEGER,
          TIME_TO INTEGER,
          REST TEXT,
          FOREIGN KEY (FAMILY_ID) REFERENCES FAMILY_ID(FAMILY_ID));''')
        cur.execute('''CREATE TABLE IF NOT EXISTS SKAMMSTAFANIR(
          LYKILL INTEGER PRIMARY KEY AUTOINCREMENT,
          SKM TEXT,
          FULL_NAME TEXT);''')
        return "Database init done."

def add_dash(year_blob):
    #"1234    1234" -> "1234 - 1234"
    years = year_blob[0] #af því að árin eru fyrsta element listanna
    x_list = []
    x = str(years)
    x = x.split(' ',1)
    x = ' -'.join(x)
    x = x.split(' ')
    x_list = [data for data in x if data not in ('', '  ', None)]
    x_list = ' '.join(x_list)
    year_blob[0] = x_list
    return year_blob


def get_fams(database_row):
    return_rows = []
    n = 0
    area_id = database_row[0]
    a = database_row[1] #nafn bæjar
    b = database_row[2] #árin
    c = database_row[3] #ábúendur
    farmnr = database_row[4] #farm_id
    area_name = database_row[5]
    d = len(b)
    for i in range(0, len(b)):
        if (b[0] in ('', None, ' ','  ')
            and c[0] in ('', None, ' ','  ')
            and n == 0): #báðar fyrstu línur tómar í blobbum, og ekkert fyrir ofan
            del b[0]
            del c[0]
        elif (b[n] in ('', None, ' ','  ')
              and c[n] in ('', None, ' ','  ')
              and n > 0): #báðar tómar, gögn í línum fyrir ofan
            return_rows.append([area_id, a, b[:n], c[:n], farmnr, area_name])
            b = b[n:] #taka út það sem búið er að bæta við sem fjölla
            c = c[n:]
            n = 0 #byrja nýja fjölluskráningu
        else:
            n +=1 #gögn á línunni, halda áfram að skrolla niður
    return return_rows


def get_farm_col(list):
    """
    Takes a list of farms of an area and returns the indexes (col nr) of the farms and their names in tuples
    """
    data_list = []
    for index, elem in enumerate(list):
        data_list.append((index, elem))
    data_list = [data for data in data_list if data[1] not in (' ', '  ', None)]
    return data_list

def create_Skammstafanir():
    #Create table Skammstafanir from Exceldocument Skammstafanir

    wb2 = load_workbook(filename='Skammstafanir-2.xlsx')

    data = wb2.worksheets[0]

    data1=data2=[0] #two rows, skm + value
    i=j=1
    data1 = [data.cell(row=i, column=1).value for i in range(1, data.max_row)]
    #print(data1)
    data2 = [data.cell(row=j, column=2).value for j in range(1, data.max_row)]
    resulting=[] #= [a.append(b) for a, b in zip(data1, data2)]
    i=0
    d1 = json.dumps(data1, default=datetime_handler)
    d2 = json.dumps(data2, default=datetime_handler)
    resulting = list(zip(data1, data2))

    while len(resulting)>0:
        insert_result = resulting[0]
        with con:
            cur.execute('INSERT INTO SKAMMSTAFANIR (SKM, FULL_NAME) VALUES (?,?)', insert_result)
        del resulting[0]
    return "skammstafanir ok"

def populate_persons(family):
    i=j=0
    for i in range(0, len(family[3])):
       a = family[0]
       b = family[1]
       c = family[2]
       d = json.dumps(family[3][j], default=datetime_handler)
       e = json.dumps(family[4][j], default=datetime_handler)
       f = family[5]
       g = family[6]
       first_person = [str(a),
                        str(b),
                        str(c),
                        str(d),
                        str(e),
                        str(f),
                        str(g)]
       with con:
            cur.execute('INSERT INTO PERSON_ID (FAMILY_ID, AREA_ID, FARM_NAME, '
                        'FAMILY_YEAR, FAMILY_DATA, FARM_ID, AREA_NAME)'
                        'VALUES (?,?,?,?,?,?,?)', first_person)
       j += 1
    return


def main():
    n = 0
    id = 1
    wb = load_workbook(filename='Vinnuskjal1.xlsx')
    e_id = 1
    ws = wb.worksheets[n]
    area_one = []
    full_list=[]
    sheet_names = wb.get_sheet_names()
    farm_counter = 1
    a = init_tables()
    print(a)
    create_Skammstafanir()
    farm_counter = 1
    while len(sheet_names) > 1:
        data = [ws.cell(row=1, column=i).value for i in
                range(ws.min_column, ws.max_column + 2)]
        new_list = get_farm_col(data)
        data_cleaned = [item for item in data if item not in (None, ' ', '  ')]
        area_id = ws.title

        with con:
            cur.execute('INSERT INTO AREA_ID (AREA_NAME) VALUES (?)', (sheet_names[0],))

            # Data to be stored in table farm_names:
        while len(data_cleaned) > 0:
            b = new_list[0][0]
            year_blob = [ws.cell(row=i, column=b).value for i in range(2, ws.max_row)]
            name_blob = [ws.cell(row=i, column=b+1).value for i in range(2, ws.max_row)]
            sql_insert = [ws.title, id, data_cleaned[0]]

            full_list.append([id, data_cleaned[0], year_blob, name_blob, farm_counter, ws.title])

            with con:
                cur.execute('INSERT INTO FARM_ID (AREA_NAME, AREA_ID, FARM_NAME) '
                            'VALUES (?,?,?)', sql_insert)
            del data_cleaned[0]
            del new_list[0]
            farm_counter += 1
        n += 1
        id += 1
        ws = wb.worksheets[n]
        sheet_names.pop(0)

    while len(full_list)>0:
        all_fams = get_fams(full_list[0])
        for i in range(0, len(all_fams)):
            dashed_years = add_dash(all_fams[i][2])
            j2 = j3 = e2 = e3 = {}
            j2 = json.dumps(dashed_years, default=datetime_handler) #family_year x2, one per family, one per einstakling
            e2 = dashed_years
            j3 = json.dumps(all_fams[i][3], default=datetime_handler) #family_data

            #e_fam should contain one value in all spots except e2 and i3 lists of persons. Those needs a row each in db
            e_fam = [e_id, #family_id
                     all_fams[i][0], #area_id
                     all_fams[i][1], #farm_id
                     e2,            #years_col + dash
                     all_fams[i][3], #name_col
                     all_fams[i][4], #farm_id
                     all_fams[i][5]] #area_name
            #print(e_fam)
            populate_persons(e_fam)
            first_fam = [str(all_fams[i][0]), str(all_fams[i][1]),
                         str(j2), str(j3), str(all_fams[i][4]),
                         str(all_fams[i][5])]
            with con:
                cur.execute('INSERT INTO FAMILY_ID (AREA_ID, FARM_NAME, '
                            'FAMILY_YEAR, FAMILY_DATA, FARM_ID, AREA_NAME)'
                            'VALUES (?,?,?,?,?,?)', first_fam)
            e_id += 1
        del full_list[0]


main()
